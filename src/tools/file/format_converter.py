"""
MCP Floating Ball - 文件格式转换工具

提供各种文件格式之间的转换功能，包括文档、图片、音频等格式。
"""

import os
import time
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
import tempfile

try:
    from PIL import Image, ImageOps
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    import docx
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import PyPDF2
    from PyPDF2 import PdfReader, PdfWriter
    PYPDF2_AVAILABLE = True
except ImportError:
    PYPDF2_AVAILABLE = False

try:
    import pypandoc
    PYPANDOC_AVAILABLE = True
except ImportError:
    PYPANDOC_AVAILABLE = False

try:
    import moviepy.editor as mp
    MOVIEPY_AVAILABLE = True
except ImportError:
    MOVIEPY_AVAILABLE = False

from ...core.logging import get_logger
from ...core.exceptions import ToolError
from ..base import BaseTool, ToolMetadata, ToolCategory, ParameterType, ToolParameter

logger = get_logger(__name__)


class FormatConverterTool(BaseTool):
    """文件格式转换工具"""

    def __init__(self):
        """初始化文件格式转换工具"""
        super().__init__()
        self.logger = get_logger("tool.format_converter")

        # 支持的转换格式
        self.supported_conversions = self._get_supported_conversions()

    def get_metadata(self) -> ToolMetadata:
        """获取工具元数据"""
        return ToolMetadata(
            name="format_converter",
            display_name="文件格式转换",
            description="文件格式转换工具，支持文档、图片、音频等多种格式转换",
            category=ToolCategory.FILE,
            version="1.0.0",
            author="MCP Floating Ball",
            tags=["file", "converter", "format", "document", "image", "audio", "video"],
            parameters=[
                ToolParameter(
                    name="input_file",
                    type=ParameterType.STRING,
                    description="输入文件路径",
                    required=True
                ),
                ToolParameter(
                    name="output_file",
                    type=ParameterType.STRING,
                    description="输出文件路径",
                    required=True
                ),
                ToolParameter(
                    name="target_format",
                    type=ParameterType.STRING,
                    description="目标格式",
                    required=False
                ),
                ToolParameter(
                    name="quality",
                    type=ParameterType.INTEGER,
                    description="转换质量（1-100）",
                    required=False,
                    default=90
                )
            ],
            examples=["将JPG图片转换为PNG", "将PDF转换为Word文档", "将MP4转换为GIF动图"]
        )

    def get_parameters(self) -> List[ToolParameter]:
        """获取工具参数定义"""
        return [
            ToolParameter(
                name="input_file",
                type=ParameterType.STRING.value,
                description="输入文件路径",
                required=True,
                examples=["/path/to/file.txt", "C:\\Users\\file.docx"]
            ),
            ToolParameter(
                name="output_file",
                type=ParameterType.STRING.value,
                description="输出文件路径",
                required=False,
                examples=["/path/to/output.pdf", "C:\\Users\\converted.jpg"]
            ),
            ToolParameter(
                name="target_format",
                type=ParameterType.STRING.value,
                description="目标格式（当output_file未指定时使用）",
                required=False,
                choices=[
                    "pdf", "docx", "txt", "html", "rtf",  # 文档
                    "jpg", "jpeg", "png", "gif", "bmp", "webp",  # 图片
                    "mp4", "avi", "mov", "mp3", "wav"  # 媒体
                ],
                examples=["pdf", "jpg", "docx"]
            ),
            ToolParameter(
                name="options",
                type=ParameterType.OBJECT.value,
                description="转换选项（如图片质量、PDF页面大小等）",
                required=False,
                examples={
                    "quality": 95,
                    "resize": [800, 600],
                    "pdf_options": {"page_size": "A4"}
                }
            )
        ]

    def execute(self, **kwargs) -> Dict[str, Any]:
        """
        执行文件格式转换

        Args:
            input_file: 输入文件路径
            output_file: 输出文件路径
            target_format: 目标格式
            options: 转换选项

        Returns:
            转换结果
        """
        try:
            input_file = kwargs.get("input_file", "")
            output_file = kwargs.get("output_file", "")
            target_format = kwargs.get("target_format", "")
            options = kwargs.get("options", {})

            if not input_file:
                raise ToolError("输入文件路径不能为空")

            if not os.path.exists(input_file):
                raise ToolError(f"输入文件不存在: {input_file}")

            if not output_file and not target_format:
                raise ToolError("必须指定输出文件路径或目标格式")

            self.logger.info(
                "开始文件格式转换",
                input_file=input_file,
                output_file=output_file,
                target_format=target_format
            )

            start_time = time.time()

            # 确定输出文件路径
            if not output_file:
                output_file = self._generate_output_path(input_file, target_format)

            # 执行转换
            result = self._convert_file(input_file, output_file, options)

            execution_time = time.time() - start_time

            self.logger.info(
                "文件格式转换完成",
                input_file=input_file,
                output_file=output_file,
                success=result["success"],
                execution_time=execution_time
            )

            return {
                "success": result["success"],
                "input_file": input_file,
                "output_file": output_file,
                "input_size": result.get("input_size"),
                "output_size": result.get("output_size"),
                "conversion_type": result.get("conversion_type"),
                "message": result.get("message", ""),
                "error": result.get("error"),
                "execution_time": execution_time
            }

        except Exception as e:
            error_msg = f"文件格式转换失败: {e}"
            self.logger.error(error_msg)
            return {
                "success": False,
                "error": error_msg,
                "execution_time": 0
            }

    def _convert_file(
        self,
        input_file: str,
        output_file: str,
        options: Dict[str, Any]
    ) -> Dict[str, Any]:
        """执行文件转换"""

        input_path = Path(input_file)
        output_path = Path(output_file)

        # 确保输出目录存在
        output_path.parent.mkdir(parents=True, exist_ok=True)

        input_size = input_path.stat().st_size if input_path.exists() else 0

        # 根据文件类型选择转换方法
        input_ext = input_path.suffix.lower()
        output_ext = output_path.suffix.lower()

        conversion_type = f"{input_ext}->{output_ext}"

        try:
            # 文档转换
            if input_ext in ['.docx', '.doc', '.txt', '.html', '.rtf'] or output_ext == '.pdf':
                return self._convert_document(input_file, output_file, options, conversion_type, input_size)

            # 图片转换
            elif input_ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff']:
                return self._convert_image(input_file, output_file, options, conversion_type, input_size)

            # 音频/视频转换
            elif input_ext in ['.mp3', '.wav', '.mp4', '.avi', '.mov', '.flv']:
                return self._convert_media(input_file, output_file, options, conversion_type, input_size)

            # Excel转换
            elif input_ext in ['.xlsx', '.xls']:
                return self._convert_excel(input_file, output_file, options, conversion_type, input_size)

            else:
                # 尝试使用pypandoc进行通用转换
                if PYPANDOC_AVAILABLE:
                    return self._convert_with_pandoc(input_file, output_file, options, conversion_type, input_size)
                else:
                    raise ToolError(f"不支持的文件转换类型: {conversion_type}")

        except Exception as e:
            return {
                "success": False,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "error": str(e),
                "message": f"转换失败: {e}"
            }

    def _convert_document(
        self,
        input_file: str,
        output_file: str,
        options: Dict[str, Any],
        conversion_type: str,
        input_size: int
    ) -> Dict[str, Any]:
        """转换文档格式"""
        input_path = Path(input_file)
        output_path = Path(output_file)
        input_ext = input_path.suffix.lower()
        output_ext = output_path.suffix.lower()

        # 使用pypandoc进行文档转换
        if PYPANDOC_AVAILABLE:
            try:
                extra_args = []
                if output_ext == '.pdf':
                    pdf_options = options.get('pdf_options', {})
                    if 'page_size' in pdf_options:
                        extra_args.extend(['-V', f'geometry:margin=1in'])
                        if pdf_options['page_size'] == 'A4':
                            extra_args.extend(['-V', 'geometry:a4paper'])

                pypandoc.convert_file(
                    input_file,
                    output_ext.lstrip('.'),
                    outputfile=output_file,
                    extra_args=extra_args
                )

                output_size = output_path.stat().st_size if output_path.exists() else 0

                return {
                    "success": True,
                    "conversion_type": conversion_type,
                    "input_size": input_size,
                    "output_size": output_size,
                    "message": f"文档转换成功: {input_ext} -> {output_ext}"
                }
            except Exception as e:
                # pypandoc失败，尝试其他方法
                pass

        # 特殊处理某些转换
        if input_ext == '.txt' and output_ext == '.pdf':
            return self._txt_to_pdf(input_file, output_file, options, conversion_type, input_size)
        elif input_ext in ['.docx', '.doc'] and output_ext == '.txt':
            return self._docx_to_txt(input_file, output_file, options, conversion_type, input_size)

        raise ToolError(f"不支持的文档转换: {conversion_type}")

    def _convert_image(
        self,
        input_file: str,
        output_file: str,
        options: Dict[str, Any],
        conversion_type: str,
        input_size: int
    ) -> Dict[str, Any]:
        """转换图片格式"""
        if not PIL_AVAILABLE:
            raise ToolError("PIL/Pillow模块不可用，请安装: pip install pillow")

        try:
            img = Image.open(input_file)

            # 处理图片旋转
            if hasattr(img, '_getexif'):
                img = ImageOps.exif_transpose(img)

            # 调整大小
            if 'resize' in options:
                size = options['resize']
                if isinstance(size, (list, tuple)) and len(size) == 2:
                    img = img.resize(size, Image.Resampling.LANCZOS)

            # 转换为RGB模式（某些格式需要）
            if output_file.lower().endswith('.jpg') or output_file.lower().endswith('.jpeg'):
                if img.mode in ('RGBA', 'P'):
                    img = img.convert('RGB')

            # 保存图片
            save_options = {}
            quality = options.get('quality', 95)
            if output_file.lower().endswith(('.jpg', '.jpeg')):
                save_options['quality'] = quality
                save_options['optimize'] = True
            elif output_file.lower().endswith('.png'):
                save_options['optimize'] = True

            img.save(output_file, **save_options)

            output_size = Path(output_file).stat().st_size

            return {
                "success": True,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "output_size": output_size,
                "message": f"图片转换成功: {conversion_type}"
            }
        except Exception as e:
            return {
                "success": False,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "error": str(e),
                "message": f"图片转换失败: {e}"
            }

    def _convert_media(
        self,
        input_file: str,
        output_file: str,
        options: Dict[str, Any],
        conversion_type: str,
        input_size: int
    ) -> Dict[str, Any]:
        """转换音频/视频格式"""
        if not MOVIEPY_AVAILABLE:
            raise ToolError("moviepy模块不可用，请安装: pip install moviepy")

        try:
            input_path = Path(input_file)
            output_path = Path(output_file)
            input_ext = input_path.suffix.lower()
            output_ext = output_path.suffix.lower()

            # 视频转换
            if input_ext in ['.mp4', '.avi', '.mov']:
                clip = mp.VideoFileClip(input_file)

                if output_ext in ['.mp3', '.wav']:
                    # 视频转音频
                    if output_ext == '.mp3':
                        clip.audio.write_audiofile(output_file)
                    else:
                        clip.audio.write_audiofile(output_file, codec='pcm_s16le')
                else:
                    # 视频格式转换
                    clip.write_videofile(output_file)

                clip.close()

            # 音频转换
            elif input_ext in ['.mp3', '.wav']:
                from moviepy.audio.io.AudioFileClip import AudioFileClip
                audio = AudioFileClip(input_file)

                if output_ext == '.mp3':
                    audio.write_audiofile(output_file)
                elif output_ext == '.wav':
                    audio.write_audiofile(output_file, codec='pcm_s16le')

                audio.close()

            output_size = output_path.stat().st_size if output_path.exists() else 0

            return {
                "success": True,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "output_size": output_size,
                "message": f"媒体文件转换成功: {conversion_type}"
            }
        except Exception as e:
            return {
                "success": False,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "error": str(e),
                "message": f"媒体文件转换失败: {e}"
            }

    def _convert_excel(
        self,
        input_file: str,
        output_file: str,
        options: Dict[str, Any],
        conversion_type: str,
        input_size: int
    ) -> Dict[str, Any]:
        """转换Excel文件"""
        if not OPENPYXL_AVAILABLE:
            raise ToolError("openpyxl模块不可用，请安装: pip install openpyxl")

        output_path = Path(output_file)
        output_ext = output_path.suffix.lower()

        try:
            if output_ext == '.csv':
                # Excel转CSV
                wb = openpyxl.load_workbook(input_file)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    csv_file = str(output_path.parent / f"{output_path.stem}_{sheet_name}.csv")
                    with open(csv_file, 'w', encoding='utf-8') as f:
                        for row in ws.iter_rows(values_only=True):
                            f.write(','.join(str(cell) if cell is not None else '' for cell in row) + '\n')
            else:
                # 其他Excel转换使用pypandoc
                return self._convert_with_pandoc(input_file, output_file, options, conversion_type, input_size)

            output_size = output_path.stat().st_size if output_path.exists() else 0

            return {
                "success": True,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "output_size": output_size,
                "message": f"Excel转换成功: {conversion_type}"
            }
        except Exception as e:
            return {
                "success": False,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "error": str(e),
                "message": f"Excel转换失败: {e}"
            }

    def _convert_with_pandoc(
        self,
        input_file: str,
        output_file: str,
        options: Dict[str, Any],
        conversion_type: str,
        input_size: int
    ) -> Dict[str, Any]:
        """使用pandoc进行通用转换"""
        if not PYPANDOC_AVAILABLE:
            raise ToolError("pypandoc模块不可用，请安装: pip install pypandoc")

        try:
            output_path = Path(output_file)
            output_ext = output_path.suffix.lstrip('.')

            pypandoc.convert_file(
                input_file,
                output_ext,
                outputfile=output_file
            )

            output_size = output_path.stat().st_size if output_path.exists() else 0

            return {
                "success": True,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "output_size": output_size,
                "message": f"文件转换成功: {conversion_type}"
            }
        except Exception as e:
            return {
                "success": False,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "error": str(e),
                "message": f"文件转换失败: {e}"
            }

    def _txt_to_pdf(
        self,
        input_file: str,
        output_file: str,
        options: Dict[str, Any],
        conversion_type: str,
        input_size: int
    ) -> Dict[str, Any]:
        """文本文件转PDF"""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter

            if not PYPDF2_AVAILABLE:
                raise ToolError("PyPDF2模块不可用，请安装: pip install pypdf2")

            page_size = options.get('page_size', 'letter')
            font_size = options.get('font_size', 12)

            # 创建PDF
            c = canvas.Canvas(output_file, pagesize=letter)
            width, height = letter

            # 读取文本文件
            with open(input_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            # 绘制文本
            y_position = height - 40
            for line in lines:
                if y_position < 40:
                    c.showPage()
                    y_position = height - 40

                c.drawString(40, y_position, line.strip())
                y_position -= font_size + 2

            c.save()

            output_size = Path(output_file).stat().st_size

            return {
                "success": True,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "output_size": output_size,
                "message": f"文本转PDF成功: {conversion_type}"
            }
        except Exception as e:
            return {
                "success": False,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "error": str(e),
                "message": f"文本转PDF失败: {e}"
            }

    def _docx_to_txt(
        self,
        input_file: str,
        output_file: str,
        options: Dict[str, Any],
        conversion_type: str,
        input_size: int
    ) -> Dict[str, Any]:
        """Word文档转文本"""
        if not DOCX_AVAILABLE:
            raise ToolError("python-docx模块不可用，请安装: pip install python-docx")

        try:
            doc = Document(input_file)
            text_content = []

            for paragraph in doc.paragraphs:
                text_content.append(paragraph.text)

            # 写入文本文件
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(text_content))

            output_size = Path(output_file).stat().st_size

            return {
                "success": True,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "output_size": output_size,
                "message": f"Word转文本成功: {conversion_type}"
            }
        except Exception as e:
            return {
                "success": False,
                "conversion_type": conversion_type,
                "input_size": input_size,
                "error": str(e),
                "message": f"Word转文本失败: {e}"
            }

    def _generate_output_path(self, input_file: str, target_format: str) -> str:
        """生成输出文件路径"""
        input_path = Path(input_file)
        output_name = f"{input_path.stem}.{target_format.lstrip('.')}"
        return str(input_path.parent / output_name)

    def _get_supported_conversions(self) -> Dict[str, List[str]]:
        """获取支持的转换格式"""
        conversions = {}

        # 文档转换
        conversions["document"] = [
            "docx->pdf", "doc->pdf", "txt->pdf", "html->pdf",
            "docx->txt", "doc->txt", "html->txt", "rtf->txt",
            "txt->docx", "html->docx"
        ]

        # 图片转换
        if PIL_AVAILABLE:
            conversions["image"] = [
                "jpg->png", "png->jpg", "gif->jpg", "bmp->png",
                "tiff->jpg", "webp->png", "jpg->webp"
            ]

        # 音频/视频转换
        if MOVIEPY_AVAILABLE:
            conversions["media"] = [
                "mp4->mp3", "avi->mp3", "mov->mp3",
                "mp3->wav", "wav->mp3"
            ]

        # Excel转换
        if OPENPYXL_AVAILABLE:
            conversions["excel"] = [
                "xlsx->csv", "xls->csv"
            ]

        return conversions

    def get_supported_conversions(self) -> Dict[str, Any]:
        """
        获取支持的转换格式列表

        Returns:
            支持的转换格式
        """
        return {
            "conversions": self.supported_conversions,
            "dependencies": {
                "pil": PIL_AVAILABLE,
                "docx": DOCX_AVAILABLE,
                "openpyxl": OPENPYXL_AVAILABLE,
                "pypdf2": PYPDF2_AVAILABLE,
                "pypandoc": PYPANDOC_AVAILABLE,
                "moviepy": MOVIEPY_AVAILABLE
            }
        }

    def batch_convert(
        self,
        input_files: List[str],
        target_format: str,
        output_dir: Optional[str] = None,
        options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        批量转换文件

        Args:
            input_files: 输入文件列表
            target_format: 目标格式
            output_dir: 输出目录
            options: 转换选项

        Returns:
            批量转换结果
        """
        results = []
        success_count = 0

        for input_file in input_files:
            try:
                # 生成输出文件路径
                if output_dir:
                    output_file = str(Path(output_dir) / f"{Path(input_file).stem}.{target_format.lstrip('.')}")
                else:
                    output_file = self._generate_output_path(input_file, target_format)

                # 执行转换
                result = self._convert_file(input_file, output_file, options or {})

                results.append({
                    "input_file": input_file,
                    "output_file": output_file,
                    "success": result["success"],
                    "error": result.get("error")
                })

                if result["success"]:
                    success_count += 1

            except Exception as e:
                results.append({
                    "input_file": input_file,
                    "output_file": "",
                    "success": False,
                    "error": str(e)
                })

        return {
            "success": success_count > 0,
            "total_files": len(input_files),
            "success_count": success_count,
            "failure_count": len(input_files) - success_count,
            "results": results
        }


# 注册工具
from ..registry import tool_registry
tool_registry.register(FormatConverterTool())